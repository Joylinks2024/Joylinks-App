import openpyxl
from openpyxl.styles import Font, PatternFill


async def export_to_excel(data, headings, filepath):
    """
    Exports data from PostgreSQL to an Excel spreadsheet using psycopg2.

    Arguments:
    connection - an open psycopg2 (this function does not close the connection)
    query_string - SQL to get data
    headings - list of strings to use as column headings
    filepath - path and filename of the Excel file

    psycopg2 and file handling errors bubble up to calling code.
    """

    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.row_dimensions[1].font = Font(bold=True, strikethrough=True)
    sheet.auto_filter.ref = "A1:J999"
    font = Font(bold=True, size=11)
    fill1 = PatternFill("solid", fgColor="FFFF99")
    # Width
    sheet.column_dimensions["A"].width = 5
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["C"].width = 15
    sheet.column_dimensions["D"].width = 25
    sheet.column_dimensions["E"].width = 20
    sheet.column_dimensions["F"].width = 13
    sheet.column_dimensions["G"].width = 20
    sheet.column_dimensions["H"].width = 13
    sheet.column_dimensions["K"].width = 10
    sheet.column_dimensions["I"].width = 8
    sheet.column_dimensions["J"].width = 12

    for rowno in ["A", "B", "C", "D", "E", "F", "G", "H", "K", "I", "J"]:
        for colno, heading in enumerate(headings, start=1):
            # Fill
            sheet[f"{rowno}1"].fill = fill1
            # Font
            sheet[f"{rowno}1"].font = font
            sheet.cell(row=1, column=colno).value = heading
    # Spreadsheet row and column indexes start at 1,
    # so we use "start = 1" in enumerate, so
    # we don't need to add 1 to the indexes.
    # This time we use "start = 2" to skip the heading row.
    for rowno, row in enumerate(data, start=2):
        for colno, cell_value in enumerate(row, start=1):
            sheet[f'A{rowno}'].font = Font(bold=True)
            if row["is_active"] == True:
                sheet[f'K{rowno}'].style = "Good"
            else:
                sheet[f'K{rowno}'].style = "Bad"
            if row['total_score'] is None:
                sheet[f'I{rowno}'].style = "Note"
            elif int(row['total_score']) >= 60:
                sheet[f'I{rowno}'].style = "Good"
            else:
                sheet[f'I{rowno}'].style = "Bad"
            sheet.cell(row=rowno, column=colno).value = cell_value

    wb.save(filepath)
